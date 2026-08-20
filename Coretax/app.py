import streamlit as st
import pandas as pd
import xml.etree.ElementTree as ET
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from io import BytesIO
import re
import traceback
from datetime import datetime

# =====================
# CONFIG
# =====================
st.set_page_config(
    page_title="Smart Coretax Error Analyzer",
    layout="wide"
)

st.title("🤖 Smart Coretax Error Analyzer")
st.caption("Mapping XML Line → Excel Cell untuk semua baris data")

# =====================
# FUNGSI BANTU
# =====================
def find_excel_column_for_field(xml_field, error_message, headers):
    """Cari kolom Excel yang sesuai dengan field XML atau error message"""
    if not headers:
        return None
    
    error_lower = error_message.lower() if error_message else ""
    
    # Mapping berdasarkan field XML (prioritas tinggi)
    field_mapping = {
        'IDPlaceOfBusinessActivity': ['id tku', 'tku', 'tempat usaha', 'id tempat'],
        'CounterpartTin': ['npwp', 'nik', 'lawan transaksi', 'counterpart'],
        'TIN': ['npwp pemotong', 'tin', 'pemotong'],
        'CounterpartPosition': ['posisi', 'jabatan', 'position'],
        'SalaryPensionJhtTht': ['gaji', 'salary', 'penghasilan'],
        'TaxExemptOpt': ['status', 'ptkp', 'tk/', 'k/'],
        'CounterpartOpt': ['wni', 'wna', 'resident'],
        'TaxObjectCode': ['kode objek', 'objek pajak'],
        'NumberOfMonths': ['bulan', 'months'],
        'InsurancePaidByEmployer': ['asuransi', 'insurance'],
        'PensionContributionJhtThtFee': ['jaminan', 'jht', 'pension'],
        'Article21IncomeTax': ['pajak', 'pph21', 'income tax'],
        'WithholdingDate': ['tanggal', 'date'],
        'WorkForSecondEmployer': ['pemberi kerja kedua', 'second employer'],
    }
    
    # 1. Coba berdasarkan field XML
    if xml_field in field_mapping:
        search_terms = field_mapping[xml_field]
        for col_idx, header in headers.items():
            header_lower = str(header).lower()
            if any(term in header_lower for term in search_terms):
                return col_idx
    
    # 2. Coba berdasarkan error message
    search_patterns = [
        ('id tempat usaha', ['id tku', 'tku', 'tempat usaha']),
        ('npwp', ['npwp', 'npwp ']),
        ('nik', ['npwp', 'nik', 'nik/npwp']),
        ('nik/npwp', ['npwp', 'nik']),
        ('gaji', ['gaji', 'penghasilan']),
        ('posisi', ['jabatan', 'posisi']),
        ('status ptkp', ['status', 'ptkp']),
        ('kode objek', ['kode objek']),
        ('asuransi', ['asuransi']),
        ('jaminan', ['jaminan', 'jht']),
        ('pajak', ['pajak', 'pph21']),
        ('tanggal', ['tanggal']),
    ]
    
    for error_pattern, excel_patterns in search_patterns:
        if error_pattern in error_lower:
            for col_idx, header in headers.items():
                header_lower = str(header).lower()
                if any(pattern in header_lower for pattern in excel_patterns):
                    return col_idx
    
    # 3. Coba cari berdasarkan kesamaan nama field (case insensitive)
    xml_field_lower = xml_field.lower()
    for col_idx, header in headers.items():
        header_lower = str(header).lower()
        # Cek apakah ada kemiripan
        if xml_field_lower in header_lower or header_lower in xml_field_lower:
            return col_idx
    
    return None

def get_employee_and_excel_row(xml_line):
    """Tentukan pegawai keberapa dan baris Excel berdasarkan line XML"""
    # Line 3 khusus (TIN)
    if xml_line == 3:
        return 0, 1  # Pegawai 0, Excel row 1
    
    # Untuk data pegawai (dimulai dari line 6)
    if xml_line >= 6:
        # Rumus: setiap pegawai = 29 lines di XML
        employee_index = ((xml_line - 6) // 29) + 1
        excel_row = 3 + employee_index  # Karena data mulai baris 4
        
        return employee_index, excel_row
    
    return None, None

def parse_xml_line_content(line_content):
    """Parse konten line XML untuk dapatkan tag dan value"""
    if not line_content:
        return None, None
    
    # Pattern untuk <tag>value</tag>
    pattern1 = re.search(r'<([^>/]+)>([^<]+)</\1>', line_content)
    if pattern1:
        return pattern1.group(1), pattern1.group(2)
    
    # Pattern untuk <tag>value</tag> dengan spasi
    pattern2 = re.search(r'<([^>]+)>\s*([^<]+)\s*</\1>', line_content)
    if pattern2:
        return pattern2.group(1), pattern2.group(2).strip()
    
    # Pattern untuk self-closing tag: <tag/>
    pattern3 = re.search(r'<([^>/]+)/>', line_content)
    if pattern3:
        return pattern3.group(1), ""
    
    # Pattern untuk opening tag saja: <tag>
    pattern4 = re.search(r'<([^>/]+)>', line_content)
    if pattern4 and '</' not in line_content:
        return pattern4.group(1), ""
    
    return None, None

# =====================
# UPLOAD FILE
# =====================
col1, col2, col3 = st.columns(3)

with col1:
    excel_file = st.file_uploader("📂 Upload Excel BPA1", type="xlsx")

with col2:
    xml_file = st.file_uploader("📂 Upload XML hasil convert", type="xml")

with col3:
    error_file = st.file_uploader("📂 Upload File Error Coretax", type=["xlsx", "txt", "csv"])

st.divider()

# =====================
# BUTTON PROSES
# =====================
if st.button("🚀 Proses Smart Detection", type="primary"):

    if not excel_file or not xml_file or not error_file:
        st.error("Semua file wajib diupload!")
        st.stop()

    try:
        # =====================
        # 1. BACA EXCEL BPA1
        # =====================
        wb = load_workbook(excel_file, data_only=True)
        
        # Cari sheet DATA
        ws = None
        for sheet_name in ["DATA", "Sheet1", "BPA1"] + wb.sheetnames:
            if sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                st.write(f"✅ **Sheet ditemukan:** {sheet_name}")
                break
        
        if not ws:
            st.error("Tidak ditemukan sheet DATA!")
            st.stop()
        
        # Header di baris 3 (asumsi standar BPA1)
        header_row = 3
        data_start_row = 4  # Data mulai baris 4
        
        # Ambil semua headers
        headers = {}
        for col in range(1, ws.max_column + 1):
            header = ws.cell(row=header_row, column=col).value
            if header:
                headers[col] = str(header).strip()
        
        st.write(f"📊 **Total kolom di Excel:** {len(headers)} (Baris {header_row})")
        st.write(f"📈 **Total baris data:** {ws.max_row - data_start_row + 1}")
        
        # =====================
        # 2. BACA XML
        # =====================
        xml_content = xml_file.read().decode('utf-8')
        xml_lines = xml_content.split('\n')
        
        st.write(f"📄 **Total lines XML:** {len(xml_lines)}")
        
        # =====================
        # 3. BACA FILE ERROR
        # =====================
        # Baca file error
        if error_file.name.endswith('.xlsx'):
            err_df = pd.read_excel(error_file)
        elif error_file.name.endswith('.csv'):
            err_df = pd.read_csv(error_file)
        else:
            # Parse file text
            content = error_file.read().decode('utf-8')
            errors = []
            lines = content.split('\n')
            
            # Coba berbagai format
            for line in lines:
                line = line.strip()
                if not line:
                    continue
                
                # Format dengan pipe
                if '|' in line:
                    parts = [p.strip() for p in line.split('|')]
                    if len(parts) >= 2:
                        try:
                            line_no = int(parts[0])
                            keterangan = parts[-1] if len(parts) >= 4 else line
                            errors.append({'line': line_no, 'keterangan': keterangan})
                        except:
                            pass
                
                # Format dengan "Line X:"
                elif 'line' in line.lower():
                    match = re.search(r'line\s*[:]?\s*(\d+)', line.lower())
                    if match:
                        line_no = int(match.group(1))
                        errors.append({'line': line_no, 'keterangan': line})
                
                # Format angka di awal
                else:
                    parts = line.split()
                    if parts and parts[0].isdigit():
                        try:
                            line_no = int(parts[0])
                            keterangan = ' '.join(parts[1:]) if len(parts) > 1 else line
                            errors.append({'line': line_no, 'keterangan': keterangan})
                        except:
                            pass
            
            err_df = pd.DataFrame(errors)
        
        if err_df.empty:
            st.warning("File error kosong atau tidak bisa dibaca!")
            st.stop()
        
        # Normalize column names
        err_df.columns = [str(c).lower().strip() for c in err_df.columns]
        
        # Cari kolom line dan keterangan
        line_col = None
        ket_col = None
        
        for col in err_df.columns:
            col_lower = col.lower()
            if any(word in col_lower for word in ['line', 'baris', 'no.']):
                line_col = col
            elif any(word in col_lower for word in ['keterangan', 'error', 'pesan', 'message', 'deskripsi']):
                ket_col = col
        
        if not line_col:
            # Coba kolom pertama yang numeric
            for col in err_df.columns:
                try:
                    pd.to_numeric(err_df[col], errors='coerce')
                    line_col = col
                    break
                except:
                    pass
        
        if not line_col and len(err_df.columns) > 0:
            line_col = err_df.columns[0]
        
        if not ket_col:
            for col in err_df.columns:
                if col != line_col:
                    ket_col = col
                    break
        
        if not ket_col:
            ket_col = line_col
        
        # Konversi line ke numeric
        err_df[line_col] = pd.to_numeric(err_df[line_col], errors='coerce')
        err_df = err_df.dropna(subset=[line_col])
        
        # Konversi ke integer
        err_df[line_col] = err_df[line_col].astype(int)
        
        st.write(f"🎯 **Kolom error:** Line='{line_col}', Keterangan='{ket_col}'")
        st.write(f"📋 **Total error:** {len(err_df)} baris")
        
        # =====================
        # 4. PROSES MAPPING ERROR
        # =====================
        hasil = []
        
        progress_bar = st.progress(0)
        total_errors = len(err_df)
        
        for idx, (_, row) in enumerate(err_df.iterrows()):
            xml_line = int(row[line_col])
            keterangan = str(row[ket_col])
            
            # Update progress
            progress_bar.progress((idx + 1) / total_errors)
            
            # Tentukan baris Excel berdasarkan XML line
            employee_index, excel_row = get_employee_and_excel_row(xml_line)
            
            if employee_index is None or excel_row is None:
                continue
            
            # Cek apakah baris Excel valid
            if excel_row > ws.max_row:
                continue
            
            # Ambil konten line XML
            if xml_line - 1 < len(xml_lines):
                line_content = xml_lines[xml_line - 1].strip()
            else:
                line_content = ""
            
            # Parse XML tag dan value
            xml_field, xml_value = parse_xml_line_content(line_content)
            
            # Jika tidak bisa parse dari line, coba cari field berdasarkan error message
            if not xml_field:
                error_lower = keterangan.lower()
                if 'id tempat usaha' in error_lower or 'tku' in error_lower:
                    xml_field = 'IDPlaceOfBusinessActivity'
                elif 'npwp' in error_lower or 'nik' in error_lower:
                    xml_field = 'CounterpartTin'
                elif 'npwp pemotong' in error_lower:
                    xml_field = 'TIN'
                elif 'gaji' in error_lower:
                    xml_field = 'SalaryPensionJhtTht'
                elif 'posisi' in error_lower:
                    xml_field = 'CounterpartPosition'
                else:
                    xml_field = 'Unknown'
            
            # Cari kolom Excel yang sesuai
            excel_col = find_excel_column_for_field(xml_field, keterangan, headers)
            
            if excel_col:
                excel_cell = f"{get_column_letter(excel_col)}{excel_row}"
                cell_value = ws.cell(row=excel_row, column=excel_col).value
                
                hasil.append({
                    'line_no': xml_line,
                    'cell_excel': excel_cell,
                    'nilai_cell': str(cell_value) if cell_value is not None else '',
                    'keterangan': keterangan,
                    'excel_row': excel_row,
                    'xml_field': xml_field,
                    'pegawai_ke': employee_index if employee_index > 0 else 'TIN'
                })
        
        progress_bar.empty()
        
        # =====================
        # 5. TAMPILKAN HASIL
        # =====================
        if hasil:
            hasil_df = pd.DataFrame(hasil)
            
            # Urutkan berdasarkan line_no
            hasil_df = hasil_df.sort_values('line_no')
            
            st.subheader("✅ Hasil Mapping Error")
            
            # Tampilkan statistik
            col1, col2, col3, col4 = st.columns(4)
            with col1:
                st.metric("Total Cell Error", len(hasil))
            with col2:
                st.metric("Jenis Error", hasil_df['keterangan'].nunique())
            with col3:
                st.metric("Total Pegawai Bermasalah", hasil_df['pegawai_ke'].nunique())
            with col4:
                st.metric("Total Baris Data di Excel", ws.max_row - data_start_row + 1)
            
            # Tampilkan hasil
            st.dataframe(
                hasil_df[['line_no', 'cell_excel', 'nilai_cell', 'keterangan', 'pegawai_ke']],
                use_container_width=True,
                hide_index=True,
                column_config={
                    "line_no": st.column_config.NumberColumn("Line No XML", width="small"),
                    "cell_excel": "Cell Excel",
                    "nilai_cell": "Nilai di Excel",
                    "keterangan": "Keterangan Error",
                    "pegawai_ke": "Pegawai Ke"
                }
            )
            
            # =====================
            # 6. DOWNLOAD HASIL (EXCEL)
            # =====================
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            
            # Buat file Excel
            output = BytesIO()
            
            with pd.ExcelWriter(output, engine='openpyxl') as writer:
                # Sheet 1: Hasil Error
                hasil_df[['line_no', 'cell_excel', 'nilai_cell', 'keterangan', 'pegawai_ke', 'xml_field', 'excel_row']].to_excel(
                    writer, index=False, sheet_name='Hasil_Error'
                )
                
                # Sheet 2: Summary
                summary_df = pd.DataFrame({
                    'Parameter': [
                        'Total Cell Error', 
                        'Jenis Error', 
                        'Pegawai Bermasalah', 
                        'Total Data Excel',
                        'Total Lines XML',
                        'File Excel',
                        'Timestamp'
                    ],
                    'Nilai': [
                        len(hasil), 
                        hasil_df['keterangan'].nunique(), 
                        hasil_df['pegawai_ke'].nunique(),
                        ws.max_row - data_start_row + 1,
                        len(xml_lines),
                        excel_file.name,
                        datetime.now().strftime("%Y-%m-%d %H:%M:%S")
                    ]
                })
                summary_df.to_excel(writer, index=False, sheet_name='Summary')
                
                # Sheet 3: Error Original
                err_display = err_df.copy()
                err_display.columns = [c.title() for c in err_display.columns]
                err_display.to_excel(writer, index=False, sheet_name='Error_Original')
                
                # Sheet 4: Data Excel Problem
                if 'excel_row' in hasil_df.columns and not hasil_df.empty:
                    problem_rows = sorted(set(hasil_df['excel_row'].astype(int).tolist()))
                    if problem_rows:
                        problem_data = []
                        for row_num in problem_rows:
                            row_dict = {'Excel_Row': row_num}
                            # Ambil beberapa kolom penting saja
                            important_cols = []
                            for col_idx, header in headers.items():
                                if any(keyword in header.lower() for keyword in ['npwp', 'nik', 'id tku', 'nama', 'gaji', 'jabatan']):
                                    important_cols.append(col_idx)
                            
                            for col_idx in important_cols[:10]:  # Maks 10 kolom
                                col_letter = get_column_letter(col_idx)
                                cell_val = ws.cell(row=row_num, column=col_idx).value
                                header = headers[col_idx]
                                row_dict[f"{col_letter}_{header[:20]}"] = cell_val
                            
                            problem_data.append(row_dict)
                        
                        if problem_data:
                            pd.DataFrame(problem_data).to_excel(writer, index=False, sheet_name='Data_Problem')
            
            output.seek(0)
            
            # Download button
            st.download_button(
                label="📥 Download Hasil (Excel)",
                data=output,
                file_name=f"coretax_error_analysis_{timestamp}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            
            # Juga sediakan opsi CSV
            csv_output = BytesIO()
            hasil_df[['line_no', 'cell_excel', 'nilai_cell', 'keterangan', 'pegawai_ke']].to_csv(csv_output, index=False)
            csv_output.seek(0)
            
            st.download_button(
                label="📥 Download Hasil (CSV)",
                data=csv_output,
                file_name=f"coretax_error_analysis_{timestamp}.csv",
                mime="text/csv"
            )
            
            st.success("✅ Proses selesai! File dapat didownload.")
            
        else:
            st.error("❌ Tidak ada hasil mapping yang ditemukan!")
            
    except Exception as e:
        st.error(f"❌ Error: {str(e)}")
        with st.expander("🔍 Detail Error"):
            st.code(traceback.format_exc())

st.divider()
st.caption("""
🔧 **Cara Kerja:**
- XML Line 3 → Excel C1 (NPWP Pemotong)
- Setiap 29 lines XML = 1 baris Excel (mulai baris 4)
- Format: XML Line → Excel Row = 3 + ((Line - 6) // 29) + 1
""")
# python -m streamlit run app.py
